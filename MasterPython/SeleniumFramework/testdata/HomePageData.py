import openpyxl


class HomePageData:

    # if this is selected in the fixture then we just use the data from here
    # each data entry in the list is a full new dictionary
    # @pytest.fixture(params=HomePageData.test_home_page_data)

    test_home_page_data = [{"Firstname": "Bert", "Surname": "Simpson", "Gender": "Male"},
                           {"Firstname": "Joy", "Surname": "Poe", "Gender": "Female"},
                           {"Firstname": "Sarah", "Surname": "Baker", "Gender": "Female"},
                           {"Firstname": "Claire", "Surname": "Tully", "Gender": "Female"},
                           {"Firstname": "Jenifer", "Surname": "Smith", "Gender": "Female"},
                           ]

    # if this is selected in the fixture then we use the data from the xl spreadsheet
    # @pytest.fixture(params=HomePageData.getTestData("Testcase2"))

    @staticmethod  # so we don't have to create an object beforehand
    def getTestData(test_case_name):  # we don't need the self keyword as the method is now static
        book = openpyxl.load_workbook(r"c:\Working Files\demo.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):  # rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]   # add the dictionary to a list which is required to match test_home_page_data

